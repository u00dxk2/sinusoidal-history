"""
Verify whether public/data/us_tfp_growth.csv is the 5-year centered rolling
average of Fernald's `dtfp` column (raw TFP growth) or `dtfp_util` column
(utilization-adjusted TFP growth).

Round-2 fact-check flagged a possible mismatch: source.md says we use `dtfp`,
but series.json and /methods describe the series as "utilization-adjusted."
This script settles which column the local CSV was derived from and whether
the centering (vs trailing) is correct.

Run from repo root:
    python scripts/verify_tfp.py
"""

from __future__ import annotations

import csv
import io
import statistics
import urllib.request
from pathlib import Path

import openpyxl

FERNALD_URL = "https://www.frbsf.org/wp-content/uploads/quarterly_tfp.xlsx"
LOCAL_CSV = Path("public/data/us_tfp_growth.csv")


def fetch_xlsx() -> bytes:
    print(f"Downloading {FERNALD_URL} ...")
    return urllib.request.urlopen(FERNALD_URL).read()


def read_annual(xlsx_bytes: bytes) -> dict[str, dict[int, float]]:
    """Return {column_name: {year: value}} for dtfp and dtfp_util."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["annual"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(headers) if h}
    out: dict[str, dict[int, float]] = {"dtfp": {}, "dtfp_util": {}}
    for r in range(2, ws.max_row + 1):
        year = ws.cell(row=r, column=col["date"] + 1).value
        if not isinstance(year, int):
            continue
        for k in ("dtfp", "dtfp_util"):
            v = ws.cell(row=r, column=col[k] + 1).value
            if isinstance(v, (int, float)):
                out[k][int(year)] = float(v)
    return out


def centered_rolling(annual: dict[int, float], window: int = 5) -> dict[int, float]:
    """5-year centered rolling mean with edge clipping (the window is shrunk
    at the boundaries rather than dropped)."""
    half = window // 2
    years = sorted(annual)
    out: dict[int, float] = {}
    for y in years:
        vals = [annual[yy] for yy in range(y - half, y + half + 1) if yy in annual]
        if vals:
            out[y] = statistics.mean(vals)
    return out


def trailing_rolling(annual: dict[int, float], window: int = 5) -> dict[int, float]:
    years = sorted(annual)
    out: dict[int, float] = {}
    for y in years:
        vals = [annual[yy] for yy in range(y - window + 1, y + 1) if yy in annual]
        if len(vals) == window:
            out[y] = statistics.mean(vals)
    return out


def read_local() -> dict[int, float]:
    out: dict[int, float] = {}
    with LOCAL_CSV.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            out[int(row[0])] = float(row[1])
    return out


def compare(label: str, computed: dict[int, float], local: dict[int, float]) -> None:
    diffs = []
    for y in sorted(set(computed) & set(local)):
        d = computed[y] - local[y]
        diffs.append(abs(d))
    if not diffs:
        print(f"  [{label}] no overlap")
        return
    max_d = max(diffs)
    mean_d = sum(diffs) / len(diffs)
    n_close = sum(1 for d in diffs if d < 0.01)
    print(f"  [{label}] {len(diffs)} overlap years; max |d| = {max_d:.4f}; mean |d| = {mean_d:.4f}; within 0.01 = {n_close}/{len(diffs)}")


def main() -> None:
    xlsx = fetch_xlsx()
    sheets = read_annual(xlsx)
    local = read_local()

    print(f"Local CSV: {len(local)} rows ({min(local)}–{max(local)})")
    print(f"Fernald dtfp: {len(sheets['dtfp'])} rows ({min(sheets['dtfp'])}–{max(sheets['dtfp'])})")
    print(f"Fernald dtfp_util: {len(sheets['dtfp_util'])} rows")
    print()

    print("Comparing local CSV against derived series:")
    for col in ("dtfp", "dtfp_util"):
        for kind, fn in (("centered-5", centered_rolling), ("trailing-5", trailing_rolling)):
            derived = fn(sheets[col], window=5)
            compare(f"{col} {kind}", derived, local)
    print()

    print("Sample year-by-year comparison (1973–1980), local vs each candidate:")
    for col in ("dtfp", "dtfp_util"):
        cd = centered_rolling(sheets[col], 5)
        td = trailing_rolling(sheets[col], 5)
        print(f"  {col}:")
        print(f"    year   local  ctr-5   tr-5")
        for y in range(1973, 1981):
            l = local.get(y, float("nan"))
            c = cd.get(y, float("nan"))
            t = td.get(y, float("nan"))
            print(f"    {y}: {l:6.3f}  {c:6.3f}  {t:6.3f}")


if __name__ == "__main__":
    main()
