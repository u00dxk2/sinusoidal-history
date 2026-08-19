"""
Build public/data/us_tfp_growth.csv from Fernald's quarterly_tfp.xlsx.

The annual sheet has two TFP-growth columns of interest:
  - `dtfp`       : raw TFP growth (no utilization adjustment)
  - `dtfp_util`  : utilization-adjusted TFP growth

This project's documentation describes the series as "utilization-adjusted,"
which corresponds to `dtfp_util`. An earlier version of this CSV was
silently derived from `dtfp` (raw); this script re-derives it from
`dtfp_util` to match the documented intent. The transformation is a 5-year
centered rolling mean with edge clipping (the window shrinks at boundaries
rather than dropping rows).

Also writes public/data/us_tfp_growth_annual.csv — the UNSMOOTHED annual
`dtfp_util` values. The spectral-verdict pipeline must run inference on this
file, never the rolled one: the 5-yr MA barely attenuates the 54–55y band
(|H| ≈ 0.99) but reddens the spectrum and corrupts the AR(1) null fit
(docs/specs/spectral-verdict-build-spec.md, data prerequisite #1). The
rolled CSV remains the display series.

Run from repo root:
    python scripts/build_us_tfp_growth.py
"""

from __future__ import annotations

import csv
import io
import statistics
import urllib.request
from pathlib import Path

import openpyxl

FERNALD_URL = "https://www.frbsf.org/wp-content/uploads/quarterly_tfp.xlsx"
OUTPUT = Path("public/data/us_tfp_growth.csv")
OUTPUT_ANNUAL = Path("public/data/us_tfp_growth_annual.csv")
COLUMN = "dtfp_util"  # utilization-adjusted TFP growth, percent
WINDOW = 5  # years, centered


def fetch_xlsx() -> bytes:
    print(f"Downloading {FERNALD_URL} ...")
    # Explicit User-Agent, same precaution as the OWID build scripts
    # (OWID began 403-ing the default urllib UA on 2026-08-18).
    req = urllib.request.Request(
        FERNALD_URL, headers={"User-Agent": "sinusoidal-history-build/1.0"}
    )
    return urllib.request.urlopen(req).read()


def read_annual(xlsx_bytes: bytes, column: str) -> dict[int, float]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["annual"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    out: dict[int, float] = {}
    for r in range(2, ws.max_row + 1):
        year = ws.cell(row=r, column=col_idx["date"]).value
        if not isinstance(year, int):
            continue
        v = ws.cell(row=r, column=col_idx[column]).value
        if isinstance(v, (int, float)):
            out[int(year)] = float(v)
    return out


def centered_rolling(annual: dict[int, float], window: int) -> dict[int, float]:
    half = window // 2
    out: dict[int, float] = {}
    for y in sorted(annual):
        vals = [annual[yy] for yy in range(y - half, y + half + 1) if yy in annual]
        if vals:
            out[y] = statistics.mean(vals)
    return out


def main() -> None:
    xlsx = fetch_xlsx()
    annual = read_annual(xlsx, COLUMN)
    rolled = centered_rolling(annual, WINDOW)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, lineterminator="\n")
        writer.writerow(["year", "tfp_growth_5yr_avg_pct"])
        for year in sorted(rolled):
            writer.writerow([year, f"{rolled[year]:.4f}"])

    with OUTPUT_ANNUAL.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, lineterminator="\n")
        writer.writerow(["year", "dtfp_util_pct"])
        for year in sorted(annual):
            writer.writerow([year, f"{annual[year]:.4f}"])

    print(f"Wrote {len(rolled)} rows to {OUTPUT}")
    print(f"Wrote {len(annual)} rows to {OUTPUT_ANNUAL} (unsmoothed)")
    print(f"Source column: {COLUMN} (utilization-adjusted TFP growth)")
    print("Sample values:")
    for y in (1948, 1965, 1973, 1980, 1995, 2008, 2025):
        if y in rolled:
            print(f"  {y}: {rolled[y]:6.3f}")


if __name__ == "__main__":
    main()
